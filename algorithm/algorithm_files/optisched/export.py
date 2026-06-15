import pandas as pd
from typing import List, Dict
import os
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

class DataExporter:
    @staticmethod
    def export_excel(df: pd.DataFrame, output_excel: str, days: List[int], day_map: Dict[int, str], time_map: Dict[int, str]):
        if df.empty:
            print("Çözüm için dışa aktarılacak veri bulunamadı.")
            return

        os.makedirs(os.path.dirname(output_excel), exist_ok=True)

        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # 1. Ham Liste Görünümü
            drop_cols = [c for c in ['DersDetay', 'Dönem_Int'] if c in df.columns]
            df_list = df.drop(columns=drop_cols)
            df_list.to_excel(writer, sheet_name='Liste_Gorunumu', index=False)
            
            # Liste görünümü sekmesini stilize et
            liste_worksheet = writer.sheets['Liste_Gorunumu']
            DataExporter._apply_basic_styling(liste_worksheet)
            
            # 2. Günlere Göre Görünüm
            has_classroom = 'Sınıf' in df.columns

            # Ders kodu + sınıf birleşimi için daha okunaklı ve detaylı görünüm
            if has_classroom:
                df = df.copy()
                df['DersDetay'] = df.apply(
                    lambda r: f"📚 {r['DersKodu']}\n🚪 Sınıf: {r['Sınıf']}" if has_classroom else f"📚 {r['DersKodu']}",
                    axis=1
                )

            for d in days:
                day_name = day_map[d]
                day_df = df[df['Gün'] == day_name]
                
                if day_df.empty:
                    pd.DataFrame().to_excel(writer, sheet_name=day_name)
                    continue
                    
                # Pivot
                pivot_df = day_df.pivot_table(
                    index='Saat',
                    columns='Dönem',
                    values='DersDetay',
                    aggfunc=lambda x: '\n─────────────────\n'.join(x)
                ).fillna('')
                
                saat_sirasi = [time_map[s] for s in range(len(time_map))]
                valid_rows = [r for r in saat_sirasi if r in pivot_df.index]
                
                unique_sems = sorted(day_df['Dönem_Int'].unique())
                col_order = [f"{s}. Dönem" for s in unique_sems]
                
                pivot_df = pivot_df.reindex(index=valid_rows, columns=col_order).fillna('')
                pivot_df.to_excel(writer, sheet_name=day_name)
                
                # Stil Uygulama
                worksheet = writer.sheets[day_name]
                DataExporter._apply_advanced_styling(worksheet, len(col_order))

    @staticmethod
    def _apply_basic_styling(ws):
        header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Genişlikleri ayarla
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    @staticmethod
    def _apply_advanced_styling(ws, num_cols):
        # Renk paleti
        header_fill = PatternFill(start_color="FF2C3E50", end_color="FF2C3E50", fill_type="solid") # Dark Blue
        time_fill = PatternFill(start_color="FFECF0F1", end_color="FFECF0F1", fill_type="solid") # Light Gray
        content_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid") # White
        course_fill = PatternFill(start_color="FFE8F8F5", end_color="FFE8F8F5", fill_type="solid") # Mint
        
        header_font = Font(color="FFFFFF", bold=True, size=12)
        bold_font = Font(bold=True, size=11, color="FF2C3E50")
        normal_font = Font(size=11)
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Sütun genişlikleri (Dönemler için daha geniş olmalı çünkü içlerinde kod var)
        ws.column_dimensions['A'].width = 16 # Saat Sütunu
        for i in range(2, num_cols + 2):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 30
            
        for row in ws.iter_rows():
            max_newlines = 1
            for cell in row:
                cell.border = thin_border
                # Hücre içi hizalamayı üstte tut ki uzun listelerde ortada sıkışmasın
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                
                if cell.row == 1:
                    # Başlık
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif cell.column == 1:
                    # Saat
                    cell.fill = time_fill
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    # İçerik
                    cell.font = normal_font
                    if cell.value and str(cell.value).strip():
                        cell.fill = course_fill
                        newlines = str(cell.value).count('\n')
                        if newlines > max_newlines:
                            max_newlines = newlines
                    else:
                        cell.fill = content_fill
            
            # Satır yüksekliğini dinamik ayarla (minimum 50, satır başı 15)
            if row[0].row > 1:
                ws.row_dimensions[row[0].row].height = max(50, (max_newlines + 1) * 16)
