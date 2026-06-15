"""Classroom-matrix Excel export.

Produces a room-centric timetable that mirrors the university's
``derslikMatris`` layout: rows are classrooms, columns are (day, time-slot)
pairs, each cell holds the course occupying that room at that time, and free
room/slots are marked ``BOS``.

Two input shapes are supported via small adapters, both reduced to a common
list of :class:`Placement` records:

* ``placements_from_schedule_df`` — the solver/post-processing DataFrame
  (columns ``Sinif``/``Sınıf``, ``d_idx``, ``slot_indices``/``s_idx``,
  ``DersKodu`` ...). Used by the standalone CLI.
* ``placements_from_courses`` — the API ``ScheduledCourseResponse`` list
  (``room``, ``day`` key, ``startTime``/``endTime``). Used by the backend.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Day-column abbreviations matching the sample matrix (Pzt/Sal/Çar/...).
DAY_ABBR: dict[int, str] = {0: "Pzt", 1: "Sal", 2: "Çar", 3: "Per", 4: "Cum", 5: "Cmt", 6: "Pzr"}
DAY_FULL: dict[int, str] = {
    0: "Pazartesi", 1: "Salı", 2: "Çarşamba", 3: "Perşembe",
    4: "Cuma", 5: "Cumartesi", 6: "Pazar",
}
EMPTY_CELL = "BOS"
NO_ROOM = {"", "Sinif Yok", "Sınıf Yok", "Derslik Yok", "Belirsiz", None}

_HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_ROOM_FILL = PatternFill("solid", fgColor="ECF0F1")
_ROOM_FONT = Font(bold=True, size=10, color="2C3E50")
_COURSE_FILL = PatternFill("solid", fgColor="E8F8F5")
_EMPTY_FONT = Font(size=9, color="B0B0B0")
_COURSE_FONT = Font(size=9, color="1A5276")
_THIN = Side(style="thin", color="D5DBDB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Soft pastel palette cycled per distinct department, so each department's
# courses share a recognisable colour across the matrix.
_DEPT_PALETTE = [
    "E8F8F5", "FDEBD0", "EBF5FB", "F5EEF8", "FEF9E7",
    "EAFAF1", "FDEDEC", "EAEDED", "F4ECF7", "E8F6F3",
    "FEF5E7", "EBEDEF",
]


@dataclass(frozen=True)
class Placement:
    room: str
    day: int
    slot: int
    label: str
    department: str = ""


def _slot_time_label(slot: int, config) -> str:
    return config.time_map.get(slot, f"slot{slot}")


# --------------------------------------------------------------------------
# Adapters: build Placement records from the two available data shapes
# --------------------------------------------------------------------------
def placements_from_schedule_df(schedule_df, *, label_fn=None) -> list[Placement]:
    room_col = None
    for col in ("Sinif", "Sınıf"):
        if col in schedule_df.columns:
            room_col = col
            break
    if room_col is None or schedule_df.empty:
        return []

    placements: list[Placement] = []
    for _, row in schedule_df.iterrows():
        room = row.get(room_col)
        if room in NO_ROOM:
            continue
        day = int(row["d_idx"])
        slot_indices = row.get("slot_indices")
        if isinstance(slot_indices, (list, tuple)) and len(slot_indices):
            slots = [int(x) for x in slot_indices]
        else:
            slots = [int(row["s_idx"])]
        dept = str(row.get("department_name", "") or "").strip()
        if label_fn is not None:
            label = label_fn(row)
        else:
            code = str(row.get("DersKodu", row.get("code", "")))
            name = str(row.get("name", "") or "").strip()
            # 'name' is like "Course Name-1 (P1)"; keep the human part only.
            name = name.split("-")[0].strip() if name else ""
            inst = str(row.get("instructor", "") or "").strip()
            parts = [code]
            if name and name.lower() not in code.lower():
                parts.append(name)
            if inst and inst not in ("-", "anonim"):
                parts.append(inst)
            label = "\n".join(parts)
        for slot in slots:
            placements.append(Placement(str(room), day, slot, label, dept))
    return placements


def placements_from_courses(courses, config) -> list[Placement]:
    """Build placements from API ScheduledCourseResponse objects."""
    day_to_idx = {v: k for k, v in config.DAY_KEY_MAP.items()}
    # start "HH:MM" -> slot index, from config.time_map ("HH:MM-HH:MM")
    start_to_slot = {label.split("-")[0]: slot for slot, label in config.time_map.items()}

    placements: list[Placement] = []
    for c in courses:
        room = getattr(c, "room", "") or ""
        if room in NO_ROOM:
            continue
        day = day_to_idx.get(getattr(c, "day", None))
        start = start_to_slot.get(getattr(c, "startTime", None))
        end = start_to_slot.get(getattr(c, "endTime", None))
        if day is None or start is None:
            continue
        # endTime is the boundary after the last hour; fall back to single slot.
        last = (end - 1) if (end is not None and end > start) else start
        inst = (getattr(c, "lecturer", "") or "").strip()
        code = getattr(c, "code", "")
        name = (getattr(c, "name", "") or "").strip()
        dept = (getattr(c, "department", "") or "").strip()
        parts = [code]
        if name and name.lower() not in code.lower():
            parts.append(name)
        if inst and inst not in ("-", "anonim"):
            parts.append(inst)
        label = "\n".join(parts)
        for slot in range(start, last + 1):
            placements.append(Placement(room, day, slot, label, dept))
    return placements


# --------------------------------------------------------------------------
# Matrix assembly + workbook writing
# --------------------------------------------------------------------------
def _occupancy(
    placements: list[Placement],
) -> tuple[dict[tuple[str, int, int], str], dict[tuple[str, int, int], str]]:
    """Return (label_map, dept_map) keyed by (room, day, slot)."""
    occ: dict[tuple[str, int, int], str] = {}
    dept: dict[tuple[str, int, int], str] = {}
    for p in placements:
        key = (p.room, p.day, p.slot)
        if key in occ and p.label not in occ[key]:
            occ[key] = f"{occ[key]} / {p.label}"
        else:
            occ.setdefault(key, p.label)
        dept.setdefault(key, p.department)
    return occ, dept


def _dept_color_map(placements: list[Placement]) -> dict[str, str]:
    depts = sorted({p.department for p in placements if p.department})
    return {d: _DEPT_PALETTE[i % len(_DEPT_PALETTE)] for i, d in enumerate(depts)}


def _room_order(placements: list[Placement], room_names: list[str] | None) -> list[str]:
    if room_names:
        rooms = list(dict.fromkeys(room_names))
        # append any placed room missing from the provided list
        extra = sorted({p.room for p in placements} - set(rooms))
        return rooms + extra
    return sorted({p.room for p in placements})


def _style_sheet(ws, n_cols: int) -> None:
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 13
    for c in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(c)].width = 16
    for row in ws.iter_rows():
        for cell in row:
            cell.border = _BORDER
            cell.alignment = _CENTER
            if cell.row == 1:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
            elif cell.column == 1:
                cell.fill = _ROOM_FILL
                cell.font = _ROOM_FONT
            elif cell.value == EMPTY_CELL:
                cell.font = _EMPTY_FONT
            else:
                cell.fill = _COURSE_FILL
                cell.font = _COURSE_FONT
    ws.row_dimensions[1].height = 26
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 30


def build_workbook(
    placements: list[Placement],
    config,
    *,
    room_names: list[str] | None = None,
    days: list[int] | None = None,
) -> Workbook:
    days = days if days is not None else list(config.days)
    slots = list(config.timeslots)
    rooms = _room_order(placements, room_names)
    occ, dept = _occupancy(placements)
    dept_colors = _dept_color_map(placements)

    wb = Workbook()
    wb.remove(wb.active)

    def color_cell(cell, room, d, s):
        key = (room, d, s)
        if key in dept:
            hexc = dept_colors.get(dept[key])
            if hexc:
                cell.fill = PatternFill("solid", fgColor=hexc)

    # ---- master sheet: rooms x (all day-slots) ----
    master = wb.create_sheet("derslikMatris")
    master.cell(1, 1, "Derslik")
    col = 2
    day_slot_cols: list[tuple[int, int]] = []
    for d in days:
        for s in slots:
            master.cell(1, col, f"{DAY_ABBR.get(d, '?')} {_slot_time_label(s, config)}")
            day_slot_cols.append((d, s))
            col += 1
    for r, room in enumerate(rooms, start=2):
        master.cell(r, 1, room)
        for c, (d, s) in enumerate(day_slot_cols, start=2):
            master.cell(r, c, occ.get((room, d, s), EMPTY_CELL))
    _style_sheet(master, len(day_slot_cols))
    for r, room in enumerate(rooms, start=2):
        for c, (d, s) in enumerate(day_slot_cols, start=2):
            color_cell(master.cell(r, c), room, d, s)

    # ---- per-day sheets: rooms x slots ----
    for d in days:
        ws = wb.create_sheet(DAY_FULL.get(d, f"Gun{d}"))
        ws.cell(1, 1, "Derslik")
        for c, s in enumerate(slots, start=2):
            ws.cell(1, c, f"{DAY_ABBR.get(d, '?')} {_slot_time_label(s, config)}")
        for r, room in enumerate(rooms, start=2):
            ws.cell(r, 1, room)
            for c, s in enumerate(slots, start=2):
                ws.cell(r, c, occ.get((room, d, s), EMPTY_CELL))
        _style_sheet(ws, len(slots))
        for r, room in enumerate(rooms, start=2):
            for c, s in enumerate(slots, start=2):
                color_cell(ws.cell(r, c), room, d, s)

    # ---- legend sheet: department -> colour ----
    if dept_colors:
        legend = wb.create_sheet("Bölümler")
        legend.cell(1, 1, "Bölüm")
        legend.cell(1, 2, "Renk")
        legend.column_dimensions["A"].width = 32
        legend.column_dimensions["B"].width = 12
        for cell in (legend.cell(1, 1), legend.cell(1, 2)):
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER
        for i, (dname, hexc) in enumerate(dept_colors.items(), start=2):
            legend.cell(i, 1, dname)
            swatch = legend.cell(i, 2, "")
            swatch.fill = PatternFill("solid", fgColor=hexc)
            for cc in (legend.cell(i, 1), swatch):
                cc.border = _BORDER
                cc.alignment = _CENTER

    return wb


def export_matrix_xlsx(placements, config, path, *, room_names=None, days=None) -> str:
    wb = build_workbook(placements, config, room_names=room_names, days=days)
    wb.save(path)
    return path


def matrix_bytes(placements, config, *, room_names=None, days=None) -> bytes:
    wb = build_workbook(placements, config, room_names=room_names, days=days)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
